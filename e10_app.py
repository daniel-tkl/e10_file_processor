import argparse
import re
from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side


TITLE_PATTERN = re.compile(r"E10 Status & Processed Wafer Count \[(.*?)\]")


def _to_text(value) -> str:
	if pd.isna(value):
		return ""
	return str(value).strip()


def extract_condition_map(condition_df: pd.DataFrame) -> dict[str, dict[str, str]]:
	"""Build equipment metadata map from the Condition sheet."""
	header_row_idx = None
	for idx in range(len(condition_df)):
		row_values = [_to_text(v) for v in condition_df.iloc[idx].tolist()]
		if "Customer" in row_values and "TEL S/N" in row_values:
			header_row_idx = idx
			break

	if header_row_idx is None:
		return {}

	headers = [_to_text(v) for v in condition_df.iloc[header_row_idx].tolist()]
	column_index = {name: i for i, name in enumerate(headers) if name}

	required_cols = ["Customer", "Area", "Building", "TEL S/N", "Customer ID", "Exposure Type"]
	if not all(col in column_index for col in required_cols):
		return {}

	tel_idx = column_index["TEL S/N"]

	mapping: dict[str, dict[str, str]] = {}
	for row_idx in range(header_row_idx + 1, len(condition_df)):
		row = condition_df.iloc[row_idx]
		tel_sn = _to_text(row.iloc[tel_idx])
		if not tel_sn:
			continue

		mapping[tel_sn] = {
			"Customer": _to_text(row.iloc[column_index["Customer"]]),
			"Area": _to_text(row.iloc[column_index["Area"]]),
			"Building": _to_text(row.iloc[column_index["Building"]]),
			"TEL S/N": tel_sn,
			"Customer ID": _to_text(row.iloc[column_index["Customer ID"]]),
			"Exposure Type": _to_text(row.iloc[column_index["Exposure Type"]]),
		}

	return mapping


def _find_week_row(df: pd.DataFrame, start_idx: int) -> int | None:
	for idx in range(start_idx + 1, min(start_idx + 6, len(df))):
		first_value = _to_text(df.iat[idx, 1]) if df.shape[1] > 1 else ""
		if first_value.startswith("ww"):
			return idx
	return None


def _find_global_week_row(df: pd.DataFrame) -> int | None:
	for idx in range(len(df)):
		first_value = _to_text(df.iat[idx, 1]) if df.shape[1] > 1 else ""
		if first_value.startswith("ww"):
			return idx
	return None


def extract_trend_rows(trend_df: pd.DataFrame) -> list[dict[str, object]]:
	"""Extract all metrics rows grouped by equipment title blocks from Trend sheet."""
	rows: list[dict[str, object]] = []
	row_idx = 0
	global_week_row_idx = _find_global_week_row(trend_df)

	while row_idx < len(trend_df):
		title = _to_text(trend_df.iat[row_idx, 0]) if trend_df.shape[1] > 0 else ""
		match = TITLE_PATTERN.search(title)
		if not match:
			row_idx += 1
			continue

		equipment = match.group(1).strip()
		if equipment.lower() == "average":
			row_idx += 1
			continue

		local_week_row_idx = _find_week_row(trend_df, row_idx)
		week_row_idx = local_week_row_idx if local_week_row_idx is not None else global_week_row_idx
		if week_row_idx is None:
			row_idx += 1
			continue

		week_labels: list[str] = []
		for col_idx in range(1, trend_df.shape[1]):
			label = _to_text(trend_df.iat[week_row_idx, col_idx])
			if label:
				week_labels.append(label)

		metric_idx = (local_week_row_idx + 1) if local_week_row_idx is not None else (row_idx + 2)
		while metric_idx < len(trend_df):
			metric_name = _to_text(trend_df.iat[metric_idx, 0])
			if not metric_name:
				break
			if TITLE_PATTERN.search(metric_name):
				break

			metric_record: dict[str, object] = {
				"Equipment": equipment,
				"Metric": metric_name,
				"_source_metric_row": metric_idx + 1,
				"_source_week_row": week_row_idx + 1,
			}
			for offset, week in enumerate(week_labels, start=1):
				if offset < trend_df.shape[1]:
					metric_record[week] = trend_df.iat[metric_idx, offset]

			rows.append(metric_record)
			metric_idx += 1

		row_idx = max(metric_idx, row_idx + 1)

	return rows


def build_summary_from_frames(condition_df: pd.DataFrame, trend_df: pd.DataFrame, sort_active_first: bool = True) -> pd.DataFrame:

	condition_map = extract_condition_map(condition_df)
	trend_rows = extract_trend_rows(trend_df)

	if not trend_rows:
		return pd.DataFrame()

	trend_table = pd.DataFrame(trend_rows)
	trend_table["_row_seq"] = range(len(trend_table))

	meta_columns = ["Customer", "Area", "Building", "TEL S/N", "Customer ID", "Exposure Type"]
	for col in meta_columns:
		trend_table[col] = trend_table["Equipment"].map(
			lambda eqp, c=col: condition_map.get(str(eqp), {}).get(c, "")
		)

	week_columns = [c for c in trend_table.columns if c.startswith("ww")]
	if sort_active_first and week_columns:
		week_numeric = trend_table[week_columns].apply(pd.to_numeric, errors="coerce").fillna(0)
		trend_table["_has_ww_data_row"] = week_numeric.ne(0).any(axis=1)

		equipment_key = trend_table["TEL S/N"].astype(str).str.strip()
		equipment_key = equipment_key.where(equipment_key.ne(""), trend_table["Equipment"].astype(str).str.strip())

		trend_table["_equipment_order"] = trend_table.groupby(equipment_key)["_row_seq"].transform("min")
		trend_table["_equipment_has_ww_data"] = trend_table.groupby(equipment_key)["_has_ww_data_row"].transform("max")

		trend_table = trend_table.sort_values(
			by=["_equipment_has_ww_data", "_equipment_order", "_row_seq"],
			ascending=[False, True, True],
			kind="mergesort",
		)
	else:
		trend_table = trend_table.sort_values(by=["_row_seq"], ascending=[True], kind="mergesort")
	ordered_columns = meta_columns + ["Metric"] + week_columns + ["_source_metric_row", "_source_week_row"]
	return trend_table[ordered_columns]


def build_summary(input_path: Path, sort_active_first: bool = True) -> pd.DataFrame:
	condition_df = pd.read_excel(input_path, sheet_name="Condition", header=None)
	trend_df = pd.read_excel(input_path, sheet_name="Trend", header=None)
	return build_summary_from_frames(condition_df, trend_df, sort_active_first=sort_active_first)


def _copy_cell_style(src_cell, dst_cell) -> None:
	dst_cell.font = copy(src_cell.font)
	dst_cell.fill = copy(src_cell.fill)
	dst_cell.border = copy(src_cell.border)
	dst_cell.alignment = copy(src_cell.alignment)
	dst_cell.number_format = src_cell.number_format
	dst_cell.protection = copy(src_cell.protection)


def _get_style_bundle(src_cell, style_cache: dict[int, tuple]) -> tuple:
	style_id = src_cell.style_id
	if style_id not in style_cache:
		style_cache[style_id] = (
			copy(src_cell.font),
			copy(src_cell.fill),
			copy(src_cell.border),
			copy(src_cell.alignment),
			src_cell.number_format,
			copy(src_cell.protection),
		)
	return style_cache[style_id]


def _apply_style_bundle(dst_cell, style_bundle: tuple) -> None:
	dst_cell.font = style_bundle[0]
	dst_cell.fill = style_bundle[1]
	dst_cell.border = style_bundle[2]
	dst_cell.alignment = style_bundle[3]
	dst_cell.number_format = style_bundle[4]
	dst_cell.protection = style_bundle[5]


def _replace_border_sides(
	border: Border,
	*,
	left: Side | None = None,
	right: Side | None = None,
	top: Side | None = None,
	bottom: Side | None = None,
) -> Border:
	return Border(
		left=left if left is not None else border.left,
		right=right if right is not None else border.right,
		top=top if top is not None else border.top,
		bottom=bottom if bottom is not None else border.bottom,
		diagonal=border.diagonal,
		diagonal_direction=border.diagonal_direction,
		vertical=border.vertical,
		horizontal=border.horizontal,
		diagonalUp=border.diagonalUp,
		diagonalDown=border.diagonalDown,
		outline=border.outline,
		start=border.start,
		end=border.end,
	)


def _apply_group_outline(ws, start_row: int, end_row: int, start_col: int = 1, end_col: int = 7) -> None:
	outline_side = Side(style="medium", color="000000")

	for col_idx in range(start_col, end_col + 1):
		top_cell = ws.cell(row=start_row, column=col_idx)
		top_cell.border = _replace_border_sides(top_cell.border, top=outline_side)

		bottom_cell = ws.cell(row=end_row, column=col_idx)
		bottom_cell.border = _replace_border_sides(bottom_cell.border, bottom=outline_side)

	for row_idx in range(start_row, end_row + 1):
		left_cell = ws.cell(row=row_idx, column=start_col)
		left_cell.border = _replace_border_sides(left_cell.border, left=outline_side)

		right_cell = ws.cell(row=row_idx, column=end_col)
		right_cell.border = _replace_border_sides(right_cell.border, right=outline_side)


def write_summary_with_format(input_path: Path, output_path: Path, summary_df: pd.DataFrame) -> None:
	wb_src = load_workbook(input_path)
	wb_out = Workbook()
	try:
		ws_trend = wb_src["Trend"]
		ws_out = wb_out.active
		ws_out.title = "Summary"

		if summary_df.empty:
			wb_out.save(output_path)
			return

		meta_columns = ["Customer", "Area", "Building", "TEL S/N", "Customer ID", "Exposure Type"]
		week_columns = [c for c in summary_df.columns if c.startswith("ww")]
		data_columns = meta_columns + ["Metric"] + week_columns
		week_source_columns = {week: idx + 2 for idx, week in enumerate(week_columns)}
		style_cache: dict[int, tuple] = {}

		first_week_row = int(summary_df.iloc[0]["_source_week_row"])
		header_metric_src = ws_trend.cell(row=first_week_row, column=1)

		for out_col_idx, col_name in enumerate(data_columns, start=1):
			cell = ws_out.cell(row=1, column=out_col_idx, value=col_name)
			if col_name.startswith("ww"):
				header_src = ws_trend.cell(row=first_week_row, column=week_source_columns[col_name])
				_apply_style_bundle(cell, _get_style_bundle(header_src, style_cache))
			else:
				_apply_style_bundle(cell, _get_style_bundle(header_metric_src, style_cache))

		write_columns = data_columns + ["_source_metric_row"]
		tel_col_idx = data_columns.index("TEL S/N") + 1
		group_end_col_idx = len(data_columns)
		for row_offset, row_values in enumerate(summary_df[write_columns].itertuples(index=False, name=None), start=2):
			source_metric_row = int(row_values[-1])
			data_values = row_values[:-1]
			metric_src = ws_trend.cell(row=source_metric_row, column=1)

			for out_col_idx, (col_name, value) in enumerate(zip(data_columns, data_values), start=1):
				if pd.isna(value):
					value = None

				out_cell = ws_out.cell(row=row_offset, column=out_col_idx, value=value)
				if col_name.startswith("ww"):
					src_cell = ws_trend.cell(row=source_metric_row, column=week_source_columns[col_name])
					_apply_style_bundle(out_cell, _get_style_bundle(src_cell, style_cache))
				else:
					_apply_style_bundle(out_cell, _get_style_bundle(metric_src, style_cache))

		group_start_row = None
		previous_tel_sn = None
		for row_idx in range(2, ws_out.max_row + 1):
			current_tel_sn = str(ws_out.cell(row=row_idx, column=tel_col_idx).value or "").strip()
			if not current_tel_sn:
				continue

			if group_start_row is None:
				group_start_row = row_idx
				previous_tel_sn = current_tel_sn
				continue

			if current_tel_sn != previous_tel_sn:
				_apply_group_outline(ws_out, group_start_row, row_idx - 1, start_col=1, end_col=group_end_col_idx)
				group_start_row = row_idx
				previous_tel_sn = current_tel_sn

		if group_start_row is not None:
			_apply_group_outline(ws_out, group_start_row, ws_out.max_row, start_col=1, end_col=group_end_col_idx)

		column_widths = {
			"A": 10,
			"B": 10,
			"C": 8,
			"D": 10,
			"E": 10,
			"F": 10,
			"G": 24,
		}
		for col_letter, width in column_widths.items():
			ws_out.column_dimensions[col_letter].width = width

		for idx in range(len(week_columns)):
			col_letter = ws_out.cell(row=1, column=8 + idx).column_letter
			ws_out.column_dimensions[col_letter].width = 11

		ws_out.freeze_panes = "H2"
		wb_out.save(output_path)
	finally:
		wb_out.close()
		wb_src.close()


def parse_args() -> argparse.Namespace:
	parser = argparse.ArgumentParser(
		description="Summarize E10Monitor Excel file into *_processed.xlsx output."
	)
	parser.add_argument("input_file", help="Path to input .xlsx file")
	parser.add_argument(
		"-o",
		"--output",
		help="Optional output path. Default: <input_name>_processed.xlsx",
	)
	parser.add_argument(
		"--sort-active-first",
		dest="sort_active_first",
		action="store_true",
		default=True,
		help="Put equipment with non-zero weekly (ww) data at the top (default: enabled).",
	)
	parser.add_argument(
		"--no-sort-active-first",
		dest="sort_active_first",
		action="store_false",
		help="Keep original equipment order from the source file.",
	)
	return parser.parse_args()


def main() -> None:
	args = parse_args()
	input_path = Path(args.input_file)

	if not input_path.exists():
		raise FileNotFoundError(f"Input file not found: {input_path}")
	if input_path.suffix.lower() != ".xlsx":
		raise ValueError("Input file must be an .xlsx file")

	output_path = Path(args.output) if args.output else input_path.with_name(f"{input_path.stem}_processed.xlsx")

	summary_df = build_summary(input_path, sort_active_first=args.sort_active_first)
	write_summary_with_format(input_path, output_path, summary_df)

	print(f"Input: {input_path}")
	print(f"Rows summarized: {len(summary_df)}")
	print(f"Output: {output_path}")


if __name__ == "__main__":
	main()
